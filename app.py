import streamlit as st
import struct
import io
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title='DAD Report — Valeo NB2',
    page_icon='🏭',
    layout='wide',
    initial_sidebar_state='expanded',
)

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1F4E79, #2E75B6);
        color: white; padding: 20px 28px; border-radius: 12px;
        margin-bottom: 24px;
    }
    .main-header h1 { margin: 0; font-size: 1.6rem; }
    .main-header p  { margin: 4px 0 0; opacity: 0.85; font-size: 0.9rem; }
    .stat-box {
        background: white; border: 1px solid #E0E7EF;
        border-radius: 10px; padding: 14px 18px; text-align: center;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .stat-box .val { font-size: 1.5rem; font-weight: 700; color: #1F4E79; }
    .stat-box .lbl { font-size: 0.78rem; color: #666; margin-top: 2px; }
    .ok-box   { background:#E2EFDA; border-radius:8px; padding:10px 14px; color:#375623; }
    .warn-box { background:#FFF2CC; border-radius:8px; padding:10px 14px; color:#7D6608; }
    .err-box  { background:#FCE4D6; border-radius:8px; padding:10px 14px; color:#843C0C; }
    div[data-testid="stSidebar"] { background: #F5F9FD; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
ALL_CH = [
    ('CH001','Top1 (Z1)',      '°C'),
    ('CH002','Top2 (Z2)',      '°C'),
    ('CH003','Top3 (Z3)',      '°C'),
    ('CH004','Top4 (Z4)',      '°C'),
    ('CH005','Top5 (Z5)',      '°C'),
    ('CH006','Top6 (Z6)',      '°C'),
    ('CH007','Top7 (Z7)',      '°C'),
    ('CH008','Bot1 (Z8)', '°C'),
    ('CH009','Bot2 (Z9)', '°C'),
    ('CH010','Bot3 (Z10)','°C'),
    ('CH011','Bot4 (Z11)','°C'),
    ('CH012','Bot5 (Z12)','°C'),
    ('CH013','Bot6 (Z13)','°C'),
    ('CH014','Bot7 (Z14)','°C'),
    ('CH015','O2 Exit',           'ppm'),
    ('CH016','Dryer zone1',       '°C'),
    ('CH017','Dryer zone2',       '°C'),
    ('CH018','O2 Entrance',       'ppm'),
]
DATA_OFFSET, RECORD_SIZE = 15008, 84
TOP_IDX = list(range(7))
BOT_IDX = list(range(7, 14))
O2_IDX  = [14, 17, 15, 16]

COLORS_TOP = ['#378ADD','#D85A30','#1D9E75','#7F77DD','#BA7517','#D4537E','#639922']
COLORS_BOT = ['#53B8E0','#E2864A','#2CB8A0','#9F77CC','#D0913A','#A85090','#5E9C20']
COLORS_O2  = {'CH015':'#2E75B6','CH018':'#C55A11','CH016':'#E36C09','CH017':'#974706'}

# ── Parser ────────────────────────────────────────────────────────────────────
def parse_dad(raw: bytes) -> list:
    records = []
    total = (len(raw) - DATA_OFFSET) // RECORD_SIZE
    for i in range(total):
        base = DATA_OFFSET + i * RECORD_SIZE
        hdr  = raw[base-8:base]
        yr,mo,dy,hr,mn,sc = hdr[0],hdr[1],hdr[2],hdr[3],hdr[4],hdr[5]
        try:
            ts = datetime(2000+yr, mo, dy, hr, mn, sc)
        except ValueError:
            continue
        rec = {'ts': ts}
        for ci,(ch,_,_) in enumerate(ALL_CH):
            min_v = struct.unpack_from('>h', raw, base + ci*4)[0]
            max_v = struct.unpack_from('>h', raw, base + ci*4 + 2)[0]
            if min_v != -32767:
                rec[f'{ch}_min'] = round(min_v/10, 1)
                rec[f'{ch}_max'] = round(max_v/10, 1)
                rec[f'{ch}_avg'] = round((min_v + max_v) / 20, 1)
                rec[ch] = rec[f'{ch}_min']  # default
            else:
                rec[f'{ch}_min'] = None
                rec[f'{ch}_max'] = None
                rec[f'{ch}_avg'] = None
                rec[ch] = None
        records.append(rec)
    return records

# ── Excel Builder ─────────────────────────────────────────────────────────────
def build_excel(data, dt_start, dt_end, zones_sel):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    norm_font = Font(name='Arial', size=9)
    thin      = Side(border_style='thin', color='C0C8D0')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill= PatternFill('solid', fgColor='E2EFDA')
    blue_fill = PatternFill('solid', fgColor='DEEAF1')
    red_fill  = PatternFill('solid', fgColor='FCE4D6')
    o2_fill   = PatternFill('solid', fgColor='FFF2CC')
    dryer_fill= PatternFill('solid', fgColor='F0FBF4')
    DAY_FILLS = ['EBF3FB','FFF9EC','F0FBF4','FDF2F8','F5F5F5']

    DATES = sorted(set(r['ts'].strftime('%Y/%m/%d') for r in data))
    date_fill_map = {d: PatternFill('solid', fgColor=DAY_FILLS[i%len(DAY_FILLS)])
                     for i,d in enumerate(DATES)}
    daily = defaultdict(lambda: {ch:[] for ch,_,_ in ALL_CH})
    for r in data:
        d = r['ts'].strftime('%Y/%m/%d')
        for ch,_,_ in ALL_CH:
            if r[ch] is not None: daily[d][ch].append(r[ch])

    def sfill(ch):
        if ch in ('CH015','CH018'): return o2_fill
        if ch in ('CH016','CH017'): return dryer_fill
        idx = int(ch[2:])-1
        return PatternFill('solid', fgColor='EBF3FB') if idx%2==0 else PatternFill('solid', fgColor='FFFFFF')

    def hcell(ws, col, row, val):
        c = ws.cell(row, col, val)
        c.fill=hdr_fill; c.font=hdr_font; c.border=border
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        return c

    def dcell(ws, col, row, val, fill=None, fmt=None, bold=False):
        c = ws.cell(row, col, val)
        c.fill=fill or PatternFill('solid',fgColor='FFFFFF')
        c.font=Font(name='Arial',size=9,bold=bold); c.border=border
        c.alignment=Alignment(horizontal='right' if isinstance(val,(int,float)) else 'left')
        if fmt: c.number_format=fmt
        return c

    sel_ch = [(ch,name,unit) for ch,name,unit in ALL_CH if name in zones_sel]
    wb = Workbook()
    n  = len(data)
    ds, de = 3, n+2
    report_name = f'{dt_start.strftime("%d-%m-%Y %H.%M")} to {dt_end.strftime("%d-%m-%Y %H.%M")}'

    # Raw Data
    ws = wb.active; ws.title='Raw Data'
    ws['A1'] = f'DAD Report — {report_name} | {n:,} records'
    ws.merge_cells(f'A1:{get_column_letter(3+len(sel_ch))}1')
    ws['A1'].font=Font(bold=True,name='Arial',size=12,color='1F4E79')
    ws.row_dimensions[1].height=22
    hcell(ws,1,2,'Timestamp'); hcell(ws,2,2,'Date'); hcell(ws,3,2,'Time')
    for i,(ch,name,unit) in enumerate(sel_ch):
        hcell(ws,4+i,2,f'{name}\n({ch})\n[{unit}]')
    ws.row_dimensions[2].height=48
    for ri,rec in enumerate(data,start=3):
        ts=rec['ts']; dp=ts.strftime('%Y/%m/%d'); tp=ts.strftime('%H:%M:%S')
        bg=date_fill_map.get(dp,PatternFill('solid',fgColor='FFFFFF'))
        dcell(ws,1,ri,ts.strftime('%Y/%m/%d %H:%M:%S'),bg)
        dcell(ws,2,ri,dp,bg); dcell(ws,3,ri,tp,bg)
        for i,(ch,_,_) in enumerate(sel_ch):
            c=ws.cell(ri,4+i,rec[ch]); c.fill=bg; c.font=norm_font
            c.border=border; c.alignment=Alignment(horizontal='right'); c.number_format='0.0'
    ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=13; ws.column_dimensions['C'].width=10
    for i in range(len(sel_ch)): ws.column_dimensions[get_column_letter(4+i)].width=14
    ws.freeze_panes='A3'; ws.auto_filter.ref=f'A2:{get_column_letter(3+len(sel_ch))}{n+2}'

    # Summary
    ws2=wb.create_sheet('Summary Stats')
    ws2['A1']=f'Summary — {report_name}'
    ws2.merge_cells('A1:H1'); ws2['A1'].font=Font(bold=True,name='Arial',size=13,color='1F4E79')
    ws2['A2']=f'{n:,} records | {dt_start.strftime("%d/%m/%Y %H:%M")} – {dt_end.strftime("%d/%m/%Y %H:%M")}'
    ws2['A2'].font=Font(italic=True,name='Arial',size=9,color='595959')
    for ci,h in enumerate(['Zone / Channel','Channel','Unit','Avg','Min','Max','Range','Std Dev'],1):
        hcell(ws2,ci,4,h)
    for zi,(ch,name,unit) in enumerate(sel_ch):
        r=5+zi; col_l=get_column_letter(4+list(ch2 for ch2,_,_ in ALL_CH).index(ch))
        fill=sfill(ch)
        dcell(ws2,1,r,name,fill,bold=True); dcell(ws2,2,r,ch,fill); dcell(ws2,3,r,unit,fill)
        raw_col = get_column_letter(4 + next(i for i,(c,_,_) in enumerate(sel_ch) if c==ch))
        for ci2,formula,f2,fmt in [
            (4,f"=ROUND(AVERAGE('Raw Data'!{raw_col}{ds}:{raw_col}{de}),1)",green_fill,'0.0'),
            (5,f"=ROUND(MIN('Raw Data'!{raw_col}{ds}:{raw_col}{de}),1)",blue_fill,'0.0'),
            (6,f"=ROUND(MAX('Raw Data'!{raw_col}{ds}:{raw_col}{de}),1)",red_fill,'0.0'),
            (7,f"=F{r}-E{r}",PatternFill('solid',fgColor='FFC7CE'),'0.0'),
            (8,f"=ROUND(STDEV('Raw Data'!{raw_col}{ds}:{raw_col}{de}),2)",PatternFill('solid',fgColor='FFFFFF'),'0.00'),
        ]:
            c=ws2.cell(r,ci2,formula); c.fill=f2; c.font=norm_font; c.border=border
            c.alignment=Alignment(horizontal='right'); c.number_format=fmt
    for col,w in zip('ABCDEFGH',[24,10,8,12,12,12,12,10]): ws2.column_dimensions[col].width=w

    # Daily Summary
    ws3=wb.create_sheet('Daily Summary')
    ws3['A1']=f'Daily Average — {report_name}'
    ws3.merge_cells(f'A1:{get_column_letter(3+len(DATES))}1')
    ws3['A1'].font=Font(bold=True,name='Arial',size=12,color='1F4E79')
    hcell(ws3,1,3,'Name'); hcell(ws3,2,3,'Channel'); hcell(ws3,3,3,'Unit')
    for di,d in enumerate(DATES):
        y,m,day=d.split('/'); hcell(ws3,4+di,3,f'{day}/{m}/{y}')
    for zi,(ch,name,unit) in enumerate(sel_ch):
        r=4+zi; fill=sfill(ch)
        dcell(ws3,1,r,name,fill,bold=True); dcell(ws3,2,r,ch,fill); dcell(ws3,3,r,unit,fill)
        for di,date in enumerate(DATES):
            vals=daily[date][ch]
            avg=round(sum(vals)/len(vals),1) if vals else ''
            c=ws3.cell(r,4+di,avg); c.fill=green_fill; c.font=norm_font; c.border=border
            c.alignment=Alignment(horizontal='right')
            if avg!='': c.number_format='0.0'
    for col,w in zip('ABCDE',[24,10,8,14,14]): ws3.column_dimensions[col].width=w


    # ── Temperature Chart ─────────────────────────────────────────────────────
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    COLORS_CHART = ['378ADD','D85A30','1D9E75','7F77DD','BA7517','D4537E','639922',
                    '53B8E0','E2864A','2CB8A0','9F77CC','D0913A','A85090','5E9C20']
    COLORS_O2C   = {'CH015':'2E75B6','CH016':'E36C09','CH017':'974706','CH018':'C55A11'}

    n_data = len(data)
    cats = Reference(ws, min_col=1, min_row=3, max_row=n_data+2)

    def make_line_chart(title, ch_list, y_min, y_max, colors):
        lc = LineChart(); lc.title = title; lc.style = 10
        lc.y_axis.title = 'Temperature (°C)' if (y_min is not None and y_min > 0) else 'Value'
        lc.y_axis.scaling.min = y_min; lc.y_axis.scaling.max = y_max
        lc.height = 14; lc.width = 28
        lc.legend.position = 'b'; lc.y_axis.numFmt = '0.0'
        lc.x_axis.tickLblSkip = max(1, n_data // 24)
        for i, (ch, name, unit) in enumerate(ch_list):
            zi = next(j for j,(c,_,_) in enumerate(ALL_CH) if c==ch)
            col_idx = 4 + next(j for j,(c,_,_) in enumerate(sel_ch) if c==ch)
            ref = Reference(ws, min_col=col_idx, min_row=2, max_row=n_data+2)
            lc.add_data(ref, titles_from_data=False)
            lc.series[i].title = SeriesLabel(v=name)
            lc.series[i].graphicalProperties.line.solidFill = colors[i % len(colors)]
            lc.series[i].graphicalProperties.line.width = 12000
            lc.series[i].smooth = True
        lc.set_categories(cats)
        return lc

    # แยก channel ตาม group
    top_sel  = [(ch,name,unit) for ch,name,unit in sel_ch if ch in [f'CH{i:03d}' for i in range(1,8)]]
    bot_sel  = [(ch,name,unit) for ch,name,unit in sel_ch if ch in [f'CH{i:03d}' for i in range(8,15)]]
    o2_sel   = [(ch,name,unit) for ch,name,unit in sel_ch if ch in ['CH015','CH016','CH017','CH018']]

    ws_c = wb.create_sheet('Temperature Chart')
    ws_c['A1'] = f'Temperature Timeline — {report_name}'
    ws_c['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')

    if top_sel:
        lc_top = make_line_chart('Top Zones: Top1–Top7', top_sel, 570, 635,
                                  [COLORS_CHART[i] for i in range(7)])
        ws_c.add_chart(lc_top, 'A3')

    if bot_sel:
        lc_bot = make_line_chart('Bottom Zones: Bottom 1–7', bot_sel, 570, 635,
                                  [COLORS_CHART[i] for i in range(7,14)])
        ws_c.add_chart(lc_bot, 'A35')

    if o2_sel:
        ws_o = wb.create_sheet('O2 & Dryer Chart')
        ws_o['A1'] = f'O2 & Dryer Timeline — {report_name}'
        ws_o['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')

        o2_only   = [(ch,n,u) for ch,n,u in o2_sel if ch in ('CH015','CH018')]
        dryer_only= [(ch,n,u) for ch,n,u in o2_sel if ch in ('CH016','CH017')]

        if o2_only:
            lc_o2 = make_line_chart('O2 Exit vs O2 Entrance', o2_only, None, None,
                                     [COLORS_O2C[ch] for ch,_,_ in o2_only])
            lc_o2.y_axis.title = 'O2 Level (ppm)'
            lc_o2.y_axis.scaling.min = None; lc_o2.y_axis.scaling.max = None
            ws_o.add_chart(lc_o2, 'A3')

        if dryer_only:
            lc_dry = make_line_chart('Dryer zone1 vs zone2', dryer_only, None, None,
                                      [COLORS_O2C[ch] for ch,_,_ in dryer_only])
            lc_dry.y_axis.title = 'Temperature (°C)'
            lc_dry.y_axis.scaling.min = None; lc_dry.y_axis.scaling.max = None
            ws_o.add_chart(lc_dry, 'A35')


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Chart Builder ─────────────────────────────────────────────────────────────
def build_chart(data, zones_top, zones_bot, zones_o2, compare_ranges=None):
    rows = 3 if any([zones_top, zones_bot, zones_o2]) else 1
    active_rows = [bool(zones_top), bool(zones_bot), bool(zones_o2)]
    n_rows = sum(active_rows)
    if n_rows == 0:
        return None

    row_heights = []
    if zones_top: row_heights.append(0.4 if n_rows==1 else 0.35)
    if zones_bot: row_heights.append(0.4 if n_rows==1 else 0.35)
    if zones_o2:  row_heights.append(0.3 if n_rows==1 else 0.25)

    titles = []
    if zones_top: titles.append('🌡️ Top Zones')
    if zones_bot: titles.append('🌡️ Bottom Zones')
    if zones_o2:  titles.append('🫧 O2 & Dryer')

    fig = make_subplots(
        rows=n_rows, cols=1,
        shared_xaxes=True,
        subplot_titles=titles,
        vertical_spacing=0.06,
        row_heights=row_heights,
    )

    def add_traces(ch_list, colors, group, row):
        ts = [r['ts'] for r in data]
        for i,(ch,name,unit) in enumerate(ch_list):
            color = colors[i] if i < len(colors) else COLORS_O2.get(ch,'#888')
            fig.add_trace(go.Scatter(
                x=ts, y=[r[ch] for r in data],
                name=name, mode='lines',
                line=dict(color=color, width=1.5),
                legendgroup=group,
                legendgrouptitle_text=group.title() if i==0 else None,
                hovertemplate=f'<b>{name}</b>: %{{y:.1f}} {unit}<extra></extra>',
            ), row=row, col=1)

    row_num = 1
    if zones_top:
        sel = [(ch,name,unit) for ch,name,unit in ALL_CH[:7] if name in zones_top]
        colors = [COLORS_TOP[i] for i,(_,name,_) in enumerate(ALL_CH[:7]) if name in zones_top]
        add_traces(sel, colors, 'Top Zones', row_num); row_num += 1
    if zones_bot:
        sel = [(ch,name,unit) for ch,name,unit in ALL_CH[7:14] if name in zones_bot]
        colors = [COLORS_BOT[i] for i,(_,name,_) in enumerate(ALL_CH[7:14]) if name in zones_bot]
        add_traces(sel, colors, 'Bottom Zones', row_num); row_num += 1
    if zones_o2:
        sel = [(ch,name,unit) for ch,name,unit in ALL_CH[14:] if name in zones_o2]
        colors = [COLORS_O2.get(ch,'#888') for ch,_,_ in sel]
        add_traces(sel, colors, 'O2 & Dryer', row_num)

    # Compare ranges (vertical lines)
    if compare_ranges:
        for label, color, dt_s, dt_e in compare_ranges:
            for r in range(1, row_num+1):
                fig.add_vrect(x0=dt_s, x1=dt_e, fillcolor=color,
                              opacity=0.08, line_width=0, row=r, col=1)
                fig.add_vline(x=dt_s, line=dict(color=color, width=1, dash='dot'), row=r, col=1)
                fig.add_vline(x=dt_e, line=dict(color=color, width=1, dash='dot'), row=r, col=1)

    fig.update_layout(
        height=max(400, n_rows * 300),
        template='plotly_white',
        hovermode='x unified',
        hoverlabel=dict(bgcolor='white', font_size=11, namelength=30),
        legend=dict(
            groupclick='toggleitem',
            bgcolor='rgba(255,255,255,0.95)',
            bordercolor='#C0C8D0', borderwidth=1,
            font=dict(size=11), tracegroupgap=10,
        ),
        margin=dict(l=60, r=20, t=50, b=20),
        plot_bgcolor='white',
    )
    fig.update_xaxes(
        tickformat='%d/%m\n%H:%M',
        dtick=1*3600*1000,        # ทุก 1 ชั่วโมง
        gridcolor='rgba(200,200,200,0.25)',
        tickfont=dict(size=10),
        showticklabels=True,      # แสดงทุก subplot
    )
    for r in range(1, n_rows+1):
        fig.update_yaxes(gridcolor='rgba(200,200,200,0.25)', row=r, col=1)

    return fig

# ══════════════════════════════════════════════════════════════════════════════
# MAIN UI
# ══════════════════════════════════════════════════════════════════════════════
ICON_B64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAIAAAAlC+aJAAAbWElEQVR42rV6eZhdVZXvWnvvM9yp7q05lRqTkIQEQoKMQVsGIUEaeIqNgoiILY2IBB/4Gm188vwaUdFueQ0+3mvR9tmNAyAG5EEj8JgakClhyASpJJVUqlJDarjzGfZe6/1xzr11qxL663/eSb4vubdOnbP32mv4/db6IRFBdDEAwty14GPj54V3MiAAIAADwIJf+49dfJTfOnIBDAC44EbVsAj8wKdzfVnM9ac3vmnuNgDk+N38H9jL3Jbjp2C8jPp6GoyCcORq1RHWrf0Ma6tFJCZmRgQhUKCA/58XMxMRE6MQQuCcUebtYM4wSIYAj7BVzdDEDABSyvrXVS/wA22MIahbChvtH22fec4u9cPgo9s//i0l0LatpGvXXcEYAwBCCOAjHlC7CeMYOIoLsTGklAKAXXsPvfTGzu27hw9NTBfLVd8PTWwhCYAoBDBxvGREIRCiT8AARCQw3gLVHZBrNzBFyyciKYTr2Jl0YnFny/Erej/yoRXHLu0CAK21lGLO8HHIRY9EtWBDDS9gpdRbO/bd80+PvPjaO6VSWQqQUsj4ZBkRATGyLRNx/FyBNaMjAteWzMzxQ+sf5zyBo8/MQMyG2DCCUE3Z7EdPPX7TVReuXTVgjEGsRQdi44EgES2ItuhFUsp/+NnDd/7kfs+rZlKulDLeOCLWbIl1P2QCZkRAFBztrfakI0wemY4b9oHxASLWNwiAxFDxdTKVueUrl934xU/U9oDzsx/iXBqdiyFWSt5y+/+8+6e/amnOKiU1AaJAFIAghOC5GAERxxdjnH1iawiBApGIiAgRueZhzCwEcv2iyPaRI4nIqeKvmQSiIZouel+75tN/91+v09oI0bgHBsBaFuI4v0ar/8E999917z93dLQagkAzACAyChYoiBkADXH0GKodZ2398VXxPM/zlRS2paSUKGR0CoaMV/WYWSmppBRC1PMuk8F4j/HmNDMyt2acv//JP7fk0rfecKXWRsp53j7vBIwxSqkX//T2BZffmE4nGQUxIAoGkFKG2hSKFSGQAXPZtJSSiBExsj02WEIbfezS3jNOXlMoVbe/t+/goYlypYyIZCiRsM9af3Imk3p/z8H9wyP5YjFKNVyPE6j9ISamaEeIUKp4zzx4z0dPX6e1lkLUg1Y0Oo8QQhvz3374UyLDANoQMRhmQFH2gnQq8ZPvfu3Zh/77D/7mWiGkFxgCNASaQTNqQA1oGAkgCMOPfvj067742Q+fsT6XzTblWgEFAPph2NbS9rlPX3Lh+Rt7e/qyuRY3mTFax77EQMxEYIiNIUNEzMRAUazq8Nbv32uMiXy4Hk+ifu5EJIR47qWtL7/+TiqdDg0zCGJghtAYpeTj//v7xy/vfvi3mzd8ZO2v7v4bPwgZBaEgFIyCAA2jAaEZUVrDI2MvvbH9ldffOXhwuFycBWZDBhE9r/ryG+8+/cLr7w/unp2ZDH2PyUShEqUyrgU/18+FWWuTTiVfemXrsy9tEULEXoMYBeG8LPHIky9qQwTAgARAzChEvlC++LwPN6XcMy74yx/d++AZn7j+xNXLPvyhlaWKh1IaRgOoGRgFSBkSNzU1rTt+VV93x+LFXa1tHY7j2JYFJkSg7u7e7p7u5paWltZOx3YTySSwASZmIiYiIIr2wcQcLYDiUEcOwwcfe2auzHIjFgJQSmpjtry727ItYxgQGYBBMCBrWtTRsnvfAVCJRcv6p6Zmhw9NtmaT3tik6LEsyyIGRFkJAh1qQ9qW0g/CYrlaqXrlqlfx/GKxUC3lCaBUKk8cnp2eLeQLxXKl4oWmXMwLaYOUlu0ACJ6ffaF2JJoYbevVLduNMUrJerJW9TuFwKnpwujEjFSKASP8wQAMCIjGkKUUSmmIpJKAmHSsSz+zcWg8/+7uA67rVoNg7bFLVy/t8wI/nXCz2XShVGlvaVp/8gnFcqVSLuowZOL2jkXAJuGoFcv621vSxFic7ZfKnpw8vG3HDmXZIES9xnAjdCSWSh08NDk5Nbuoo5WIozBWjbBkJl8qVrwo7cT7jzbAtepqOF/RRCAQtTFBqcgmYOYocE86/tgPrVnt+b4UwvP94UNTUojevj5gIAZtDBEHQVAuV40xzbnmbDanQ03dvVKq5vHxg4fGC/lpZAUoGGPEwABMjLUcUyyVp2fyizpaY8TLrBrBW6ni+UEU5hiVxhoyRUChiVNJ6/TVba/tOEQMALxtcORgvqoct1KtrD2mz5ZydGySmZSUCIACPENkjNGGOE4sHP3HUBDoKHBDrYWQ+fxMLpdVUkxOjgtlRQWqXlmYGQAR0feDUrnSgANRNRaFQBsDgICR2wAKBgQUgCgFFj390ZXZx2/oP+Nbh4qeTrnOmpX9Y2/trVa9O26+6pz1J1V937ZtMqSUil4mpTBElmX7vl+t+kQEiGFoQq0d2wmCoOr52lCotectobNOe2/wwEObH52dnUbBUe3EqCjUvcoYzw/qCAoQRWP5DMKowmKUPSMXoig/MwMKCHw+NAzVkhBiplD605Zdi9tbkM3KJb3trc25TMaxVDqZKBQKI6OjADx6aGx6eubZ517Ys2efbVkMMDl5uFAsAvPzL760b2i/YzsRJFPKTqeaHMfJ5lra2zp16DMRkGFiYmKoAVxiHepGUqBqxIWhVrlARJUQGwpkhBiAggBnZ0CHhkw2nfjSZecdmCwMvj9YKFWKpUqpUiXiZDK5e+/Qfb+4/+rPXT504ICUat/Q/kqlcsXll729bfsbb2zp6eldumRg4vD0n15/49JPfcpxkr4fhFoHlvF9v1IupTOZZN4tVyrKcrgBFwoEgKg8z+1ARai6zuqo5nrMEGFlZoAaWkYdwmyBQ62UmskX/9fPHoFsk7CU74cVL6h4AREDSq2pr7f3sSeeXL16FRuSluW47uThmcG9Q/liqcvoFSuOnZ7J94Q9QjqlStUY0toQY6gNMXl+tbm1zdCEHwRCRimRYztTA+KKXQjnE1umuKhxDcPXaiKgAE1QKAFRoE1nW8umL1/WnMsBY9ULCqVqoVgpFivFsjc9W1i+/JjF3T2Hxg+H2hSLJc8PDs/MImBba9vAkuWvvv7Gw5sfNoZmCwXPC0vlaqXiVaq+5wdERIa01rnmZiUFaQNkIK5tBEwLSJmazzKRGSkCVIgYFWgSwDXOoQ1UAtbGtqzRsYnfPfBHp6cXgXw/CkcGgGKpsmzZMq2N7TjVSgVAdHUtTiZTqVS6vbWdAIVUiUTyC1+4xvN8BsEMgFJH+cgQMDEREQGwAGbSiCpyZBQIEbw7SleidgpRAiViQAYhGZCZagfBoDVUfCDhh2Frc9P377h59/D4z3/1hy3b3p+YKu7ZO6i1VsoSGLMcIQQwSCW1NmGoo/PVWkfQn4jCMDDGBEFg205nV9/o6DCRIVYRLiJjmAgERd0UbKjSXHMhNZ9fc0P5q6WqqJZEaMQY8DUbZSk1PT17330PZboWgcC2lub+vp43t25NZlrSuWap7DjxIc5RFCJjtNFhBN7CUGtisBh8r1qcIOK+3h7fqwS+RyZEVIjIZIA0s8QaHa6zubrjL3Qh4HoAMxNhnUdH4W0MBBpZhlo357LXXXvFo8+/mc+PuLadcJx0MtXa0RHmx/a8+5ptW4rJRrYpcNi3kCUYC0zkIYYxMFDCZKAh1X8Ct3RQUHId17Jty3J1UA6Copts4hofYGTkiIJzA51BwDgLzVECYAYBTDyPehLF9JwYQiZDQojZfPG1V/6tJGyU0vODSsULQ82M02PDo2+/uCihshikpUkJ4yI5gm1kgYAIxOAbrJCwWJRKgW5q5eYVRmvP88NAEzOgkNIiY8hoAACmmMLzwpUBoJpH5+stgtpdHBcQjl4MRKANMIfatOSaMrnc8MFxIYXvhVUvYGatdU4EXWm3xZXr7HCZTc0252zO2OAoFAKAwTB6IRd9LoXmEWFVuTxiDBH7QaijbhMDgEAhhRBIFLEzRowOZEHzR81vSSAgMjc0X5iZWQBWvYCZMHYhIuJiscgmVEJqMlXPL1c8rU3VCzLay4kwJyEnTbulO5LclsaUC7aFQgoA0Ib9AAplKnjYPksz4DNpMibwA611jVUSABhDVAkwKYTAmDxyY4xijVLWKnEctfUYJwIiBqDAX7tqqeu6oTagTaBNLpvp6V68/c2tQkkm4/uh5wU6DL2qJ03YbGNOUZtLHWlub8L2FtnS4TR1JdM96cziZLbDbWlR7TnZnoJ2B1KsOdRkjO8HOgzj9TOTDjMJueGU9oyDTITAOLfOubAV87pazMAUGSBqDAgpgsmpazdd+blPnvtfbv1hNmGBY1tK/ufb7vrON776ic9f7k3PSCk93y8WS6HvB76HpFMKMopbHGhNiWxaJHK27ExjX7MYaIb+ZtHTZLen0k1WNiOaXXAgYO0TmdAPwiAkMswkQecPV//qxPSTX0j+5SokL5RIsQshzpm4TikbahvW6CUhsK56nf2LfvSt68/77I39hf33feXCitP88Frwn/j1Z2/63t133JLJZUwQFmZnZmemfd/TQYhsHIkJCQkbki66KQtak9yV4+5W6u7gxR3UmePWhMxYTkImbCGBwsDTYRgEfhgETMRa54v6+jMzl7XkH3h48Pk9BaGYiGsOUkujeJQYgDjqGQBBSmGmSn9x9SVbdgxuf+GFqZvOhrdemR0bawHzq1Vy3YO/m9j0xXPPPv33v/lDOpNNpbJKCN/3FOuEAkdwk6LmBLjNFnSlYWkbtLZCIgXGQH4GkgJ0qMIwZ5NFxg9DTbparWodsgmLVXPX+ZkT3clrnj68s6CqxlY2ELNAbBxRHKUS1/fHQEgAQoAJVwx0v7lr78lJhpefG52cEiinNaUxOEbn395zcFlvFwReuVQpuIWq7yvpl43MprLaolFLJRjcqq3GpTKhGC2jCskY41XMDAfjtjebHGaaYjcMQ2JTrVbDarlYCv9hQ2aJGbt081Qq6SzOJYbyuk7O6twL59Dogm43cdQurHX8aHxyeu3Ja4fKZnqyUPAhJE3aaEuNkWzLZiYmp4DJ87xCsegHoWV771RT2/O5RS5sySeTI2GKg7QpOWbKBoNsDIuSEaMhTINdgKRRjt+c0VKz0X4pv2em/J0znJU0+fmnpnKpBEhZ1BERqbGzBXEMPG8Dc9g7ao0RQyLx2B+f/9q1V870rbpz+PXr2sBocKX1oxF2Vq9b3d/zzHP/BslEuVK2bUcHwfDs7ECS2bJ3Tuf7E6IMooCWg1IKG4AlYiXUvq1uPvPEfNn7zeD4/z1UyvihZYclPyjOFL61IjiN81c+fTiTTGoQmqRBm8GPqfnC0VBMKbnOPTmGPXHDk8jIVPKdV7c+/NhTT9x/z4aLr/rj4K5Vtt7uw0Tfqsfvu/sX//Lbkfd2i9a2wPNmi4VCWL1iybJbe7KKur++fddv9oxmpItAUaqWCJVQd2bcH5+44ji2dpWrX+rvOLuz9Y493gEfL8j5f7t6+uDI8Be2zDhuyiAELAlEyFhb/RETkjkwhw3REbEajjk/A4imphtu+vZ9P/neq89t/vUfnh46MPyZgf5PnX/O448/8d2//b7V3EaMs5ViVuE9a84cn9z749e3AeuLFrefsL7v1ncGF7mOZiMAgjDMNuH9p67+x9ffvGHssGEOjL5xZdcDx7cNVs0JwmufHLv6nYK2khKlx6BBEIiQBULEzoEXWj+KgXkDosa7oogglCIk66qrN513/jkXnr9h2UnHj09OXfPlr7787AuQzoVhGXTg4sBtS859ctuzP9/7WspNEJHZvef5czack3Ef3TfYlkxXAt9H+uOffexnb+66b2h8WSYDwDbp7+yaWHZgakVC3FXR4xpYWJZlV0CGTBolAZoIHMS4BueW+YFZiAgEN45AmAgtC7ItT21+9KnNj4G0wPhgJ0W2FZBOX/Ixm/iSgbN/ueWpJ8e2dqVbLIES4TDyDW9s+ekJ5747k99XmHDsxK+P3/Da0Oj/2D/cm875KEMTMqiWhDpkzJ5ZnVS2ZamQOWQVgghBEiMBEiCDiKHoPMwzbz4wn9LUh6bMwIRCcrUsFfd8/KrJHa9UDo/IZLvRPunixnXXXn7c9W/veeKHr/1i2JtoSzRXKQwIBXJSuVsq+e/u2PLLgf/0VmlsuZN9eWTw9tG3O5NNFWZbWJqZmEIAIVVS2ELIEIVvNIMyIDQQA1J9rFbHCvNnA7VCthDiRdiVAACl4uJMqr0jd/5N9qqLu3qeD95+cGLfK03p1rPXfqs0M3P7v35lf3koZD9jN1VIA6NCkIyGOGenHyoMvbV96vRM30/Lb77jjbe6GZ8ifxAGJUdTFQYCliCJMWQBIA2DiWI3gm44HywcbUKzgILN/eXZ8eya9dYFfz3+3O/52YckiIGVf35K7wWJAHfuf2nn0CPgJKRIOCLtsam9UxCwYdAATXZ6v64MFrZZaGXdbFBngkSa6q1bYEBiJBQaEIhqhkcGigYLc4dwlBPAua+jIVOctQgkUPNF1wSrzz/8wA/g4LuQaTdB+f3RLXtTi3RQBlMSyVapXKO90AQY0wcygARgmAQIYlaokipBzAHp+sxYA2iOmDsiIgMBKAYgJohtT/HoKbIj1uahgEdLo7VLYMMonDQ2txdT3f7Pv4HVAjd3o1QsLHSMJg9tG6GNWRsgAAIEZormnxz9g8gYsymNyBD5C8ZwN+pbYg1aMnJDUz3ijRzhZ4GMHyBbwCOghFIy9n5iEFLnp/XmH2MizY6LxkOwkDUzC0QEQxQKBKAqACESA0shmBnAAJCQEX9lFCIaHcX5mqPhctSgIkAVgWSOGz4x+6t3F5ABhWTSEH2L0SKP0lZBALCVBIFM9cGAAMdh7QEwa83aACLYNgcBGALHMsZARBOZQQjj+aAkKAmBBqPBtoAZwhBsO25gEoOUEARABEqClAB+bFwiCEJIuFGPFojAtWvalChwo+BHy7I+0IWSri0EgiGcy6SAiKypua0l15TRWg+PjHf1dicTzp79BzOZdLFYUpaylOV5/tJVx4yNHy4Xiu2L2nLZpr0HRhzL6u5q37NvmImUbSdcp1goLh7ocWxntlAsFktSSCHR84NUwu3saBvcNwzMi/t7HMcaOnCImRBxTixDLC2VjDYJWJuRNeTZXCaZdCSTiVARAgMTIiCZrlz67759/dnr11oC/v62r1yy8QzIF+68bdPlnzzvh7dtuvC8M5b2L/rmps/f/s3rcmn3O7f81ZWXbhRB8OlPnLvpS5d+/forOT+14axT7/3RN7lSOu3E1T++/abFTe6Gs0479aTjrviLj7e3Nl9+ycYbr/nMtZ//pCPx7jtu/vOzThULUzsTmWTCyTZlGnGniAkOIgB0tDW3t2RZ62jpUR+PiETC2bFlx1vb3/8/T78sEBMJd+/+EUAYG5+44lPnr14+UCgWkwk3l00f3D8cBGG2KTMycigsFtOp5JL+7qrnA4iLzj+zt3vRuvWnPPzYM9t3DW57a2tbW0tLc1NHW4sQmEy6SwZ6iqWSFMJ13b37DlAYohD1oouAbHRHW0tnR+ucKAZR1aPbGJNOJdYeu2Ro+3ZhyUgVgFHQGYMpd9/QsJSodbjzvT1dna2YTr29bdcDm5/sbGuuVP1yufLA75984F82Yzq1Y+fubDaTbG/d+d7gxMRkoVRp6e3Z+vaOhx97Zmlf17vbd+3Ze0Cm2p559uWbv3rV+OT05MGDuwf3jYyOzcyWfK13vrdncfciYSlDRmA8RUcEDII1q5anUqlIN7FQbqMNWZa675e/u+bLf21lMoYYAOOcGB1XEIBSgAg6tBwn1AbKJWAAx4HQB6kAJSCBMSCVUJK8KkgFhoA0OA4EBrwS2C6mM1wqgmUBSjAh+D6kM8AMoQbXAQTQ2nIdrU1dfATASokgn//He++85urLtdZSyMixRAx/EKUUzHzxx8/qX9Zv/KpAQiCMu8EMAMKxI72IsO1QayHFwBlnta06XgjoP+3P2lesFsJ0rjyu7+SPCIFutmX5mRudbLNMpZZ+5GO57gHhyGM+dkHX2pM5qLavXtN/8oeFgLZjjl1y5gZpWaiUSCVqz3dCTQ2ll6VA7fsDy5defMF50bSvXtNEbaAKCECGOtpbN11zBVUrSkTsnrDOduK6T0wMRNKy12z8ZO8JJ5NfPfasjyezWWE7A+tOM4EHAAPrTkk0ZbuPWwcAay+6vHvVWgZsHVjWu/YU5bg9J5yyaMUqKue715y0ZuMlluMAmXqzgZkxYgEIkfMoKSjQX/3SlZ2dbQ0qRQYAedttt9WlKohojFl97Iqt23a9t22nm0owxZbAKG81tAWEwMrs5PTwvsrsjHKdamG2MHZQOgk7kZw5OBQGQaq1fWT7WzrwStMT0weHqjNTmfbO2fHR2QN7UFkmCKYPjejQmzqwt3B4jIxBISAeuDAiCACBgMC2ktVideMF53731hstWwkhcO5wMNLM1Ss1M0MY6t27h6768te3bHnHzaRrwSAaKn1N0BD4ICUqi70qKAul5MAHRLRs1iEYA1KiVOxHP1UceACAtsthAABoOxz6QISOi7G/Y0PLgRFYCvQ8c9JJ6/7prtuWr1hi2xbO15fV5TZzYENr7XnB+7uHbvn2nc+8+IpSUkrBNeEgz4kq4/oYYa2ok4eiLmuL1UvAjEJE1b0maeJ4bhD/JxKwNej4InkSoGY0YJ975vrvfePaFcsHEkk30vAdTS/Ec0QAEPwgLJcq4+NTv/zto7995F8PHBxlMrJ+dPOEiXOH18A8YvLH80zD8/lGoyHrr48FGoYRrUR/X99lF5195SXndna0pdJJx7GZ6+4+T3Z5FEWkHwTlcqVa9d8f3P/MC6+9unXb0IGRQrHoB4E2xERUE8HFc1gAEGJ+cwDndcywLmqNZ0XRgcQxhkIIKZXluG4umx3o6z71hJXnrD9h5bJeN5nIZFKObTXYeE7Ci2SoQSY7T62rDVWqnlf1fc+fnsmPT05PzxYKxbLvB6ExxlCdLdX2g411fcFRRSrHeFpal+/F22QEVEolXLspnWptbupsb27JZWzXcV03lXKVlPVnxwNhrnkxEdUwbqPaN356NJPzPD8MtTFk4nVzg4QSBGKkD2ViFDinLawpHKJOGSIKFHXcH2nV6pJNrHF2FKiUUpZybMt2bEvJBv3u/OVFXhgNNBs0pThvXIN1sQITkSHmSJFUd+/5jdW6XRuD4Ai1bWO8YLT62I8QEFEKISLVYx0HHKmjPkJ2ybUJ3xEbXdAXOKI39kEa6Hm2wH9Hl35UkXjDemBeggH4d0h9Q+eoIc6P+vA4Khe2y+b6H9igIl0gnsf5mQkbsgjP+zjPsY9cCQLA/wPoBgIMvBaiJgAAAABJRU5ErkJggg=="

st.markdown(f"""
<div class="main-header">
    <h1><img src="data:image/png;base64,{ICON_B64}" style="height:48px;vertical-align:middle;margin-right:10px;border-radius:10px"> DAD Report Generator</h1>
    <p>Valeo APU3 — NB2 Furnace Data Analysis</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('### 📂 Upload ไฟล์ .DAD')
    uploaded = st.file_uploader(
        'เลือกไฟล์ .DAD (หลายไฟล์ได้)',
        type=['DAD','dad'],
        accept_multiple_files=True,
    )

    st.markdown('---')
    st.markdown('### ⏰ ช่วงเวลา')

    from datetime import date, time
    col1, col2 = st.columns(2)
    with col1:
        d_start = st.date_input('วันเริ่ม', value=date.today())
        t_start = st.time_input('เวลาเริ่ม', value=time(7, 0))
    with col2:
        d_end = st.date_input('วันสิ้นสุด', value=date.today())
        t_end = st.time_input('เวลาสิ้นสุด', value=time(19, 0))

    dt_start = datetime.combine(d_start, t_start)
    dt_end   = datetime.combine(d_end,   t_end)

    st.markdown('---')
    st.markdown('### 📈 ค่าที่แสดงในกราฟ')
    value_mode = st.radio(
        '',
        options=['MIN', 'AVG (เฉลี่ย)', 'MAX'],
        index=0,
        horizontal=True,
        label_visibility='collapsed',
    )
    mode_key = {'MIN': '_min', 'AVG (เฉลี่ย)': '_avg', 'MAX': '_max'}[value_mode]

    st.markdown('---')
    st.markdown('### 📊 เลือก Zone ที่อยากดู')

    all_names_top = [name for _,name,_ in ALL_CH[:7]]
    all_names_bot = [name for _,name,_ in ALL_CH[7:14]]
    all_names_o2  = [name for _,name,_ in ALL_CH[14:]]

    with st.expander('🌡️ Top Zones', expanded=True):
        col_a, col_b = st.columns(2)
        sel_top = []
        for i, name in enumerate(all_names_top):
            col = col_a if i % 2 == 0 else col_b
            if col.checkbox(name, value=True, key=f'top_{i}'):
                sel_top.append(name)
    with st.expander('🌡️ Bottom Zones', expanded=True):
        col_a, col_b = st.columns(2)
        sel_bot = []
        for i, name in enumerate(all_names_bot):
            col = col_a if i % 2 == 0 else col_b
            if col.checkbox(name, value=True, key=f'bot_{i}'):
                sel_bot.append(name)
    with st.expander('🫧 O2 & Dryer', expanded=True):
        col_a, col_b = st.columns(2)
        sel_o2 = []
        for i, name in enumerate(all_names_o2):
            col = col_a if i % 2 == 0 else col_b
            if col.checkbox(name, value=True, key=f'o2_{i}'):
                sel_o2.append(name)

    st.markdown('---')
    st.markdown('### 🔀 เปรียบเทียบช่วงเวลา')
    compare_mode = st.toggle('เปิดโหมดเปรียบเทียบ')
    compare_ranges = []
    if compare_mode:
        n_compare = st.number_input('จำนวนช่วงที่เปรียบเทียบ', 1, 3, 2)
        COMP_COLORS = ['#FF6B6B','#4ECDC4','#45B7D1']
        for ci in range(int(n_compare)):
            with st.expander(f'ช่วงที่ {ci+1}', expanded=True):
                cc1, cc2 = st.columns(2)
                with cc1:
                    cd_s = st.date_input(f'วันเริ่ม', key=f'cd_s{ci}')
                    ct_s = st.time_input(f'เวลาเริ่ม', key=f'ct_s{ci}')
                with cc2:
                    cd_e = st.date_input(f'วันสิ้นสุด', key=f'cd_e{ci}')
                    ct_e = st.time_input(f'เวลาสิ้นสุด', key=f'ct_e{ci}')
                compare_ranges.append((
                    f'ช่วง {ci+1}', COMP_COLORS[ci],
                    datetime.combine(cd_s, ct_s),
                    datetime.combine(cd_e, ct_e),
                ))

    run_btn = st.button('🚀 Generate Report', type='primary', use_container_width=True)

# ── Main Area ─────────────────────────────────────────────────────────────────
if not uploaded:
    st.info('📂 กรุณา Upload ไฟล์ .DAD ในแถบซ้ายก่อนครับ')
    st.stop()

if dt_end <= dt_start:
    st.error('❌ เวลาสิ้นสุดต้องมากกว่าเวลาเริ่มต้นครับ')
    st.stop()

# Parse & Filter
@st.cache_data(show_spinner='⏳ กำลัง parse ไฟล์...')
def load_data(files_data, dt_s, dt_e, mode_key):
    all_records = []
    for fname, raw in files_data:
        recs = parse_dad(raw)
        filtered = [r for r in recs if dt_s <= r['ts'] <= dt_e]
        all_records.extend(filtered)
    all_records.sort(key=lambda x: x['ts'])
    seen = set()
    unique = []
    for r in all_records:
        k = r['ts']
        if k not in seen:
            seen.add(k)
            # set rec[ch] ตาม mode ที่เลือกก่อน cache
            new_r = dict(r)
            for ch,_,_ in ALL_CH:
                new_r[ch] = r.get(f'{ch}{mode_key}')
            unique.append(new_r)
    return unique

files_data = [(f.name, f.read()) for f in uploaded]
data = load_data(tuple((n,d) for n,d in files_data), dt_start, dt_end, mode_key)

# ── Summary Cards ──────────────────────────────────────────────────────────────
duration = dt_end - dt_start
hrs = int(duration.total_seconds() // 3600)
mins = int((duration.total_seconds() % 3600) // 60)

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f'<div class="stat-box"><div class="val">{len(data):,}</div><div class="lbl">Records</div></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="stat-box"><div class="val">{len(uploaded)}</div><div class="lbl">ไฟล์ .DAD</div></div>', unsafe_allow_html=True)
with col3:
    st.markdown(f'<div class="stat-box"><div class="val">{hrs}h {mins}m</div><div class="lbl">ระยะเวลา</div></div>', unsafe_allow_html=True)
with col4:
    days_with_data = len(set(r['ts'].strftime('%Y/%m/%d') for r in data))
    st.markdown(f'<div class="stat-box"><div class="val">{days_with_data}</div><div class="lbl">วันที่มีข้อมูล</div></div>', unsafe_allow_html=True)

st.markdown('<div style="margin-top:16px"></div>', unsafe_allow_html=True)

# ── Coverage Check ────────────────────────────────────────────────────────────
with st.expander('📋 ตรวจสอบความครบถ้วนของข้อมูล', expanded=(len(data)==0)):
    day_counts = defaultdict(int)
    for r in data:
        day_counts[r['ts'].strftime('%Y/%m/%d')] += 1

    check_day = dt_start.date()
    all_ok = True
    while check_day <= dt_end.date():
        day_str = check_day.strftime('%Y/%m/%d')
        count = day_counts.get(day_str, 0)
        if count > 0:
            st.markdown(f'<div class="ok-box">✅ {check_day.strftime("%d/%m/%Y")} — {count:,} records</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="err-box">❌ {check_day.strftime("%d/%m/%Y")} — ไม่มีข้อมูล (ยังไม่ได้ upload)</div>', unsafe_allow_html=True)
            all_ok = False
        st.markdown('<div style="margin-top:4px"></div>', unsafe_allow_html=True)
        check_day += timedelta(days=1)

if len(data) == 0:
    st.error('🚨 ไม่มีข้อมูลในช่วงเวลาที่ระบุ กรุณาตรวจสอบช่วงวัน/เวลา หรือ Upload ไฟล์เพิ่มครับ')
    st.stop()

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_chart, tab_summary, tab_compare, tab_download = st.tabs([
    '📈 Chart', '📊 Summary', '🔀 เปรียบเทียบ', '⬇️ Download Excel'
])

with tab_chart:
    fig = build_chart(data, sel_top, sel_bot, sel_o2,
                      compare_ranges if compare_mode else None)
    if fig:
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning('กรุณาเลือก Zone อย่างน้อย 1 Zone ในแถบซ้ายครับ')

with tab_summary:
    st.markdown(f'**ช่วงเวลา:** {dt_start.strftime("%d/%m/%Y %H:%M")} → {dt_end.strftime("%d/%m/%Y %H:%M")}')
    st.markdown(f'**ค่าที่แสดง:** {value_mode}')
    st.markdown(f'**Records:** {len(data):,}')
    st.markdown('---')

    all_sel_names = sel_top + sel_bot + sel_o2
    sel_channels = [(ch,name,unit) for ch,name,unit in ALL_CH if name in all_sel_names]

    import pandas as pd
    rows_summary = []
    for ch,name,unit in sel_channels:
        vals = [r[ch] for r in data if r[ch] is not None]
        if vals:
            rows_summary.append({
                'Zone / Channel': name,
                'Channel': ch,
                'Unit': unit,
                'Avg': round(sum(vals)/len(vals),1),
                'Min': round(min(vals),1),
                'Max': round(max(vals),1),
                'Range': round(max(vals)-min(vals),1),
            })
    if rows_summary:
        df = pd.DataFrame(rows_summary)
        st.dataframe(df, use_container_width=True, hide_index=True)

with tab_compare:
    if not compare_mode:
        st.info('เปิด "โหมดเปรียบเทียบ" ในแถบซ้ายก่อนครับ')
    else:
        for label, color, dt_s, dt_e in compare_ranges:
            seg = [r for r in data if dt_s <= r['ts'] <= dt_e]
            st.markdown(f'**{label}:** {dt_s.strftime("%d/%m/%Y %H:%M")} → {dt_e.strftime("%d/%m/%Y %H:%M")} ({len(seg):,} records)')
        if compare_ranges:
            st.markdown('---')
            import pandas as pd
            all_sel_names = sel_top + sel_bot + sel_o2
            sel_channels = [(ch,name,unit) for ch,name,unit in ALL_CH if name in all_sel_names]
            comp_rows = []
            for ch,name,unit in sel_channels:
                row = {'Zone': name, 'Unit': unit}
                for label,_,dt_s,dt_e in compare_ranges:
                    vals = [r[ch] for r in data if dt_s <= r['ts'] <= dt_e and r[ch] is not None]
                    row[f'{label} Avg'] = round(sum(vals)/len(vals),1) if vals else '-'
                    row[f'{label} Min'] = round(min(vals),1) if vals else '-'
                    row[f'{label} Max'] = round(max(vals),1) if vals else '-'
                comp_rows.append(row)
            st.dataframe(pd.DataFrame(comp_rows), use_container_width=True, hide_index=True)

with tab_download:
    st.markdown('### ⬇️ Download Excel Report')
    all_sel_names = sel_top + sel_bot + sel_o2
    if not all_sel_names:
        st.warning('กรุณาเลือก Zone ในแถบซ้ายก่อนครับ')
    else:
        if st.button('📊 สร้าง Excel Report', type='primary'):
            with st.spinner('⏳ กำลังสร้าง Excel...'):
                excel_data = build_excel(data, dt_start, dt_end, all_sel_names)
            fname = f'DAD Report {dt_start.strftime("%d-%m-%Y %H.%M")} to {dt_end.strftime("%d-%m-%Y %H.%M")}.xlsx'
            st.download_button(
                label='⬇️ Download Excel',
                data=excel_data,
                file_name=fname,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary',
                use_container_width=True,
            )
            st.success(f'✅ พร้อม Download: {fname}')
